import openpyxl

book = openpyxl.open("/home/arun/Documents/Test_Data.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)

sheet.cell(row=2, column=2).value = "Arun Balasubramanian"
print(sheet.cell(row=2, column=2).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet['A2'].value)

data_list = []
temp_dict = {}

for i in range(2, sheet.max_row + 1):

    for j in range(2, sheet.max_column + 1):

        print(sheet.cell(row=1, column=j).value)
        print(sheet.cell(row=i, column=j).value)
        temp_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    data_list.append(dict(temp_dict))
    temp_dict.clear()


print(data_list)
