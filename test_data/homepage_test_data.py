import openpyxl

class Homepage_test_data:
    form_test_data = [{"name": "Ram", "email": "ram@india.com", "gender": "Male"},
                      {"name": "Sita", "email": "Sita@india.com", "gender": "Female"},
                      {"name": "Lakshman", "email": "Lakshman@india.com", "gender": "Male"}]

    @staticmethod
    def get_data_excel():
        book = openpyxl.open("/home/arun/Documents/Test_Data.xlsx")
        sheet = book.active
        data_list = []
        temp_dict = {}

        for i in range(2, sheet.max_row + 1):

            for j in range(2, sheet.max_column + 1):
                print(sheet.cell(row=1, column=j).value)
                print(sheet.cell(row=i, column=j).value)
                temp_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            data_list.append(dict(temp_dict))
            temp_dict.clear()

        return data_list
